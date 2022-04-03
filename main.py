import requests
from bs4 import BeautifulSoup
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles.fonts import Font

today = datetime.datetime.today()
file_date = date = today.strftime("%Y%m%d")

keywords = ["コーヒー", "珈琲", "coffee"]

nikkei_url = "https://www.nikkei.com/search"

class NikkeiSearch:

    def __init__(self):
        self.articles = []
        self.title_list = []
        self.desc_list = []
        self.url_list = []

    def search(self, name):
        params = {"keyword": name, "volume": "10"}
        response = requests.get(nikkei_url, params=params)
        print(response.url)
        if response.status_code == requests.codes.ok:
            response_html = response.text
            soup = BeautifulSoup(response_html, "html.parser")
            titles = soup.find_all(name="h3", class_="nui-card__title")
            descriptions = soup.find_all(name="a", class_="nui-card__excerpt")
            # replace 全角("　")と半角(" ")どっちも必要。
            self.title_list = [title.text.replace("　", "").replace(" ", "").replace("\n", "") for title in titles]
            self.url_list = [desc.get_attribute_list("href")[0] for desc in descriptions]
            self.desc_list = [desc.text.replace("　", "").replace(" ", "").replace("\n", "") for desc in descriptions]
        else:
            self.title_list.append("Something wrong with the URL.")

    def format_data(self):
        for n in range(len(self.title_list)):
            article = {}
            article["title"] = self.title_list[n]
            article["description"] = self.desc_list[n]
            article["url"] = self.url_list[n]
            self.articles.append(article)

    def export_data(self):
        df = pd.DataFrame(self.articles)
        df.duplicated(subset=["title"])
        self.number_of_articles = len(df)
        df.to_excel(f"{file_date}_nikkei_coffee.xlsx", encoding='utf_8_sig', index=False)

    def format_excel(self):
        wb = openpyxl.load_workbook(f"{file_date}_nikkei_coffee.xlsx")
        ws = wb["Sheet1"]
        font = Font(name="メイリオ", size=10)
        for rows in ws["A1:C100"]:
            for cell in rows:
                cell.font = font

        ws["A1"].value = "Title"
        ws["B1"].value = "Description"
        ws["C1"].value = "URL"
        fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="ceefc2")
        heading_list = ["A1", "B1", "C1"]
        for heading in heading_list:
            ws[heading].fill = fill
            ws[heading].font = Font(name="メイリオ", size=12, bold=True)

        n = 2
        for url in range(self.number_of_articles):
            ws[f"C{n}"].hyperlink = ws[f"C{n}"].value
            n += 1
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 70
        ws.column_dimensions["c"].width = 55
        wb.save(f"{file_date}_nikkei_coffee.xlsx")


nikkei = NikkeiSearch()

for keyword in keywords:
    nikkei.search(keyword)
    nikkei.format_data()
nikkei.export_data()
nikkei.format_excel()